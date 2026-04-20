import os
import re
import csv
import json
from pathlib import Path

def parse_log_line(line: str) -> dict:
    """
    Parses a log line with the format:
      key1: value1 key2: value2 ...
    Returns a dict of {key: value}.
    """
    # Split on ': ' but be careful: values can contain 'e-05' etc.
    # Strategy: find all occurrences of "word: " to identify key positions.
    pattern = re.compile(r'(\w+):\s')
    keys = [(m.group(1), m.start()) for m in pattern.finditer(line)]

    record = {}
    for i, (key, start) in enumerate(keys):
        # Value starts after "key: "
        val_start = start + len(key) + 2
        # Value ends just before the next key (or end of line)
        val_end = keys[i + 1][1] if i + 1 < len(keys) else len(line)
        value = line[val_start:val_end].strip()
        # Try to cast to number
        try:
            value = int(value)
        except ValueError:
            try:
                value = float(value)
            except ValueError:
                pass  # keep as string
        record[key] = value

    return record


def convert_log(input_path: str, output_csv: str = None, output_xlsx: str = None):
    """
    Reads a .txt log file and converts it to CSV and/or XLSX.

    Args:
        input_path:  path to the input .txt log file
        output_csv:  path for the output .csv file (optional)
        output_xlsx: path for the output .xlsx file (optional)
    """
    input_path = Path(input_path)

    # Default output names alongside the input file
    if output_csv is None:
        output_csv = input_path.with_suffix('.csv')
    if output_xlsx is None:
        output_xlsx = input_path.with_suffix('.xlsx')

    # Parse all lines
    records = []
    with open(input_path, 'r', encoding='utf-8') as f:
        for raw_line in f:
            line = raw_line.strip()
            if not line:
                continue
            record = parse_log_line(line)
            if record:
                records.append(record)

    if not records:
        print("No records found. Check the log format.")
        return

    # Collect all column names (preserving insertion order)
    all_keys = list(dict.fromkeys(k for r in records for k in r))

    # ── CSV ────────────────────────────────────────────────────────────────────
    with open(output_csv, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=all_keys)
        writer.writeheader()
        writer.writerows(records)
    print(f"CSV saved → {output_csv}")

    # ── XLSX ───────────────────────────────────────────────────────────────────
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Logs"

        header_font  = Font(bold=True, color="FFFFFF")
        header_fill  = PatternFill("solid", fgColor="2F5597")
        header_align = Alignment(horizontal="center", vertical="center")

        # Header row
        for col_idx, key in enumerate(all_keys, start=1):
            cell = ws.cell(row=1, column=col_idx, value=key)
            cell.font  = header_font
            cell.fill  = header_fill
            cell.alignment = header_align

        # Data rows
        for row_idx, record in enumerate(records, start=2):
            for col_idx, key in enumerate(all_keys, start=1):
                ws.cell(row=row_idx, column=col_idx, value=record.get(key, ""))

        # Auto-fit column widths
        for col_idx, key in enumerate(all_keys, start=1):
            max_len = max(
                len(str(key)),
                max((len(str(r.get(key, ""))) for r in records), default=0)
            )
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 30)

        wb.save(output_xlsx)
        print(f"XLSX saved → {output_xlsx}")

    except ImportError:
        print("openpyxl not installed — skipping XLSX. Install with: pip install openpyxl")

    print(f"\nDone! {len(records)} rows, {len(all_keys)} columns.")
    print("Columns:", all_keys)


# ── Entry point ────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    usage = 'ide'  # ['console', 'ide']
    if usage == 'console':
        import sys

        if len(sys.argv) < 2:
            print("Usage: python convert_log.py <input_log.txt> [output.csv] [output.xlsx]")
            sys.exit(1)

        inp  = sys.argv[1]
        csv_ = sys.argv[2] if len(sys.argv) > 2 else None
        xls_ = sys.argv[3] if len(sys.argv) > 3 else None
    else:
        folder = '/home/user/Scrivania/PhD/DGM/docs/logs esperimenti/2026-04-07T13-13-29'
        inp = os.path.join(folder, 'log.txt')
        csv_ = os.path.join(folder, 'log_csv.csv')
        xls_ = os.path.join(folder, 'log_x.xlsx')

    convert_log(inp, csv_, xls_)