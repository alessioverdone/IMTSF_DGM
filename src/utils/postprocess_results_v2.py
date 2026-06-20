import os
import re
import csv
from pathlib import Path

import pandas as pd


def parse_log_line(line: str) -> dict:
    """
    Parses a log line with the format:
      key1: value1 key2: value2 ...
    Returns a dict of {key: value}.
    """
    pattern = re.compile(r'(\w+):\s')
    keys = [(m.group(1), m.start()) for m in pattern.finditer(line)]

    record = {}
    for i, (key, start) in enumerate(keys):
        val_start = start + len(key) + 2
        val_end = keys[i + 1][1] if i + 1 < len(keys) else len(line)
        value = line[val_start:val_end].strip()
        try:
            value = int(value)
        except ValueError:
            try:
                value = float(value)
            except ValueError:
                pass
        record[key] = value

    return record


def parse_log_file(file_path: Path) -> list[dict]:
    """Reads a single .txt log file and returns a list of parsed records."""
    records = []
    with open(file_path, 'r', encoding='utf-8') as f:
        for raw_line in f:
            line = raw_line.strip()
            if not line:
                continue
            record = parse_log_line(line)
            if record:
                records.append(record)
    return records


def convert_all_logs(root_dir: str, output_csv: str = None, output_xlsx: str = None):
    """
    Recursively finds all .txt files under `root_dir`, parses each one as a
    log file, and produces a single combined CSV + XLSX with an extra column
    `source_folder` containing the parent folder name of each log file.

    Args:
        root_dir:    path to the root directory to scan
        output_csv:  path for the combined .csv  (default: <root_dir>/all_logs.csv)
        output_xlsx: path for the combined .xlsx (default: <root_dir>/all_logs.xlsx)
    """
    root = Path(root_dir)
    if output_csv is None:
        output_csv = root / 'all_logs.csv'
    if output_xlsx is None:
        output_xlsx = root / 'all_logs.xlsx'

    # ── 1. Find all .txt files recursively ─────────────────────────────────
    txt_files = sorted(root.rglob('*log.txt'))
    if not txt_files:
        print(f"No .txt files found under {root}")
        return

    print(f"Found {len(txt_files)} .txt file(s) under {root}\n")

    # ── 2. Parse each file into a DataFrame and collect ────────────────────
    frames = []
    for txt_path in txt_files:
        records = parse_log_file(txt_path)
        if not records:
            print(f"  ⚠  {txt_path.relative_to(root)} — no records, skipped")
            continue

        df = pd.DataFrame(records)
        # Add the parent folder name as last column
        df['source_folder'] = txt_path.parent.name
        frames.append(df)
        print(f"  ✔  {txt_path.relative_to(root)} — {len(df)} rows")

    if not frames:
        print("\nNo valid records found in any file.")
        return

    # ── 3. Concatenate everything ──────────────────────────────────────────
    combined = pd.concat(frames, ignore_index=True)
    # Move source_folder to the very last column (it should already be, but
    # concat can reorder if files have different columns)
    cols = [c for c in combined.columns if c != 'source_folder'] + ['source_folder']
    combined = combined[cols]

    print(f"\nCombined: {len(combined)} rows, {len(combined.columns)} columns")

    # ── 4. Save CSV ────────────────────────────────────────────────────────
    combined.to_csv(output_csv, index=False, encoding='utf-8')
    print(f"CSV  saved → {output_csv}")

    # ── 5. Save XLSX with formatting ───────────────────────────────────────
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "Logs"

        header_font = Font(name="Arial", bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="2F5597")
        header_align = Alignment(horizontal="center", vertical="center")
        data_font = Font(name="Arial")

        all_keys = list(combined.columns)

        for col_idx, key in enumerate(all_keys, start=1):
            cell = ws.cell(row=1, column=col_idx, value=key)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        for row_idx, (_, row) in enumerate(combined.iterrows(), start=2):
            for col_idx, key in enumerate(all_keys, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=row[key])
                cell.font = data_font

        for col_idx, key in enumerate(all_keys, start=1):
            max_len = max(
                len(str(key)),
                combined[key].astype(str).str.len().max() if len(combined) else 0
            )
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 35)

        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = "A2"

        wb.save(output_xlsx)
        print(f"XLSX saved → {output_xlsx}")

    except ImportError:
        print("openpyxl not installed — skipping XLSX.")

    print(f"\nDone! {len(combined)} total rows, {len(combined.columns)} columns.")
    print("Columns:", list(combined.columns))

def print_top_n_by_dataset(csv_path: str | None, n: int = 5,
                           dataset_col: str = 'dataset_name',
                           metric_col: str = 'test_mse_mean'):
    """
    Reads the unified CSV and, for each unique dataset, prints the `n` rows
    with the smallest `metric_col` value.

    Args:
        csv_path:    path to the combined CSV produced by convert_all_logs
        n:           number of top (lowest) rows to show per dataset
        dataset_col: name of the column that identifies the dataset
        metric_col:  name of the column to rank by (ascending)
    """
    df = pd.read_csv(csv_path)

    if dataset_col not in df.columns:
        print(f"Column '{dataset_col}' not found. Available: {list(df.columns)}")
        return
    if metric_col not in df.columns:
        print(f"Column '{metric_col}' not found. Available: {list(df.columns)}")
        return

    # Make sure the metric is numeric (drop rows where it isn't)
    df[metric_col] = pd.to_numeric(df[metric_col], errors='coerce')
    df_valid = df.dropna(subset=[metric_col])

    datasets = sorted(df_valid[dataset_col].unique())
    print(f"\n{'='*80}")
    print(f"  Top {n} rows with lowest '{metric_col}' per dataset")
    print(f"{'='*80}")

    for ds in datasets:
        subset = df_valid[df_valid[dataset_col] == ds]
        top = subset.nsmallest(n, metric_col)

        print(f"\n── {ds} ({len(subset)} total rows) "
              f"── best {len(top)} ──────────────────────────")
        print(top.to_string(index=False))

    print(f"\n{'='*80}\n")





# ── Entry point ────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    usage = 'ide'  # ['console', 'ide']

    if usage == 'console':
        import sys
        if len(sys.argv) < 2:
            print("Usage: python convert_logs_recursive.py <root_dir> [output.csv] [output.xlsx]")
            sys.exit(1)
        root = sys.argv[1]
        csv_ = sys.argv[2] if len(sys.argv) > 2 else None
        xls_ = sys.argv[3] if len(sys.argv) > 3 else None
    else:
        # ── Set your root folder here ──
        root = '/home/user/Scrivania/PhD/DGM/docs/logs esperimenti/raw_logs'
        csv_ = os.path.join(root, 'all_logs_v2.csv')
        xls_ = os.path.join(root, 'all_logs_v2.xlsx')

        # Cumulate results of different runs in a single .csv and .xlsx file
        # convert_all_logs(root, csv_, xls_)

        # Print the top N best results per dataset
        print_top_n_by_dataset(csv_, n=10)